"""Parse the schedule XLSX into per-group lesson lists."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

import openpyxl

# "Дисциплина/Дисциплины по выбору" plus OCR/typo variants (e.g. "Дисөиплины").
_ELECTIVE_HEADER_RE = re.compile(r"дис\w*\s+по\s+выбору", re.IGNORECASE)
_WEBINAR_ONLY_RE = re.compile(r"\(?вебинары?\)?", re.IGNORECASE)

# Instructor pattern: Surname I.O.
_INSTR_RE = re.compile(r"([А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.[А-ЯЁ]?\.?)")

# Room numbers, optionally prefixed with "ауд." / "в".
_ROOM_RE = re.compile(
    r"(?:ауд\.?\s*)?(?:\bв\s+)?\b(\d{3,4}(?:-\d{3,4})?)\b",
    re.IGNORECASE,
)

_MONTH_NAMES: dict[str, int] = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}

# Maps Cyrillic day-of-week markers found in the spreadsheet to Python
# weekday indices (0 = Monday … 6 = Sunday).
DAY_MARKERS: dict[str, int] = {
    "ПОНЕДЕЛЬНИК": 0,
    "ВТОРНИК": 1,
    "СРЕДА": 2,
    "ЧЕТВЕРГ": 3,
    "ПЯТНИЦА": 4,
    "СУББОТА": 5,
    "ВОСКРЕСЕНЬЕ": 6,
}

# Regex that matches a day-of-week row (column A contains the day name
# surrounded by asterisks and spaces).
_DAY_RE = re.compile(
    r"\*\s*(" + "|".join(
        r"\s*".join(ch for ch in day)
        for day in DAY_MARKERS
    ) + r")\s*",
    re.IGNORECASE,
)

# Time slot pattern like "08.30-10.00" or "08:30-10:00".
_TIME_RE = re.compile(r"(\d{2})[.:](\d{2})\s*-\s*(\d{2})[.:](\d{2})")

# Packed elective lines mix several week/subgroup clauses, e.g.
# "Психология управления, … 1-9 нед. гр.№4, с 10 нед. гр.№5".
_SEGMENT_CUT_RE = re.compile(
    r"""
    [,;]\s*
    (?=с\s+\d+\s*(?:по\s+\d+\s*)?нед)
    """,
    re.IGNORECASE | re.VERBOSE,
)

_ROOM_THEN_WEEK_RE = re.compile(
    r"\b\d{3,4}(?=\s+с\s+\d+\s*(?:по\s+\d+\s*)?нед)",
    re.IGNORECASE,
)

_GROUP_MARK_RE = re.compile(
    r"(?:подгрупп[аы]|групп[аы]|гр\.?)\s*№?\s*\d+",
    re.IGNORECASE,
)

_LECTURE_WEEK_RE = re.compile(
    r"\bлек(?:ция|ции|ц)?\.?\s*\d+(?:\s*-\s*\d+)?\s*нед",
    re.IGNORECASE,
)

_COURSE_META_RE = re.compile(
    r"""
    \d+\s*-\s*\d+\s*нед\.?
    | с\s+\d+\s*(?:по\s+\d+\s*)?нед\.?
    | \d+\s*нед\.?
    | \bлек(?:ция|ции|ц)?\.?\b
    | \bпрак(?:тика)?\.?\b
    | (?:под)?групп[аы]\s*№?\s*\d+
    | \bгр\.?\s*№?\s*\d+
    | \(\s*\d+\s*гр\.?\s*\)
    | \b\d+\s*гр\.?
    | \bвебинары?\b
    | \(по выбору\)
    """,
    re.IGNORECASE | re.VERBOSE,
)


@dataclass(frozen=True)
class Lesson:
    """A single lesson entry."""

    day: int  # 0=Monday … 6=Sunday
    time_start: str  # "HH:MM"
    time_end: str  # "HH:MM"
    subject: str
    instructor: str
    room: str
    notes: str = ""
    link: str = ""
    type: str = ""  # "Лекц", "Прак" or empty
    weeks: str = "all"  # "all", "even", "odd", or a range like "1-9" / "10-18"
    start_from: str = ""  # optional "MM-DD" — skip occurrences before this date


def _normalize_day(text: str) -> str:
    """Remove extra spaces between characters in day-of-week markers."""
    return re.sub(r"\s+", "", text).upper()


def _is_elective_header(text: str) -> bool:
    return bool(_ELECTIVE_HEADER_RE.search(text or ""))


def _is_usable_subject(subject: str) -> bool:
    """True if a parsed subject is a real lesson, not a header or leftover marker."""
    subject = (subject or "").strip().rstrip(":").strip()
    if not subject or _is_elective_header(subject):
        return False
    if _WEBINAR_ONLY_RE.fullmatch(subject):
        return False
    return True


def _strip_elective_header(line: str) -> str:
    """Remove a leading 'Дисциплина(ы) по выбору:' prefix from a line."""
    return _ELECTIVE_HEADER_RE.sub("", line, count=1).lstrip(" :").strip()


def _split_elective_segments(line: str) -> list[str]:
    """Break a packed elective line into one week-range / one subgroup pieces."""
    if not (line or "").strip():
        return []
    starts = [0]
    week_clause_starts: list[int] = []
    for match in _SEGMENT_CUT_RE.finditer(line):
        starts.append(match.end())
        week_clause_starts.append(match.end())
    for match in _ROOM_THEN_WEEK_RE.finditer(line):
        starts.append(match.end())
        week_clause_starts.append(match.end())

    group_marks = list(_GROUP_MARK_RE.finditer(line))
    for index, mark in enumerate(group_marks):
        before = line[: mark.start()]
        sep = re.search(r"[,;]\s*$", before)
        if index == 0 and not _LECTURE_WEEK_RE.search(before):
            continue
        if any(ws <= mark.start() <= ws + 48 for ws in week_clause_starts):
            continue
        if sep:
            starts.append(sep.end())
        elif index > 0:
            starts.append(mark.start())

    starts = sorted(set(starts))
    parts: list[str] = []
    for i, start in enumerate(starts):
        end = starts[i + 1] if i + 1 < len(starts) else len(line)
        if i + 1 < len(starts):
            comma = _SEGMENT_CUT_RE.search(line, start, end)
            if comma:
                end = comma.start()
        part = line[start:end].strip(" ,;")
        if part:
            parts.append(part)
    return parts or [line.strip()]


def _bare_course_title(subject: str) -> str:
    """Course name with week/subgroup/lecture markers removed."""
    cleaned = _COURSE_META_RE.sub(" ", subject or "")
    cleaned = re.sub(r"[().]", " ", cleaned)
    return re.sub(r"[\s,;:]+", " ", cleaned).strip(" ,;:-")


def _subject_is_metadata(subject: str) -> bool:
    return not _bare_course_title(subject)


def _keep_group_markers_on_subject(subject: str, notes: str) -> tuple[str, str]:
    """Move week/subgroup leftovers from notes onto the subject for matching."""
    if not notes:
        return subject, notes
    if notes.casefold() in (subject or "").casefold():
        return subject, notes
    if re.search(
        r"(?:под)?групп[аы]\s*№?\s*\d+|\bгр\.?\s*№?\s*\d+|\d+\s*нед",
        notes,
        re.IGNORECASE,
    ):
        return f"{subject} {notes}".strip(), ""
    return subject, notes


def _parse_elective_line(line: str) -> list[tuple[str, str, str, str]]:
    """Parse one elective row, inheriting the course title across split segments."""
    raw_items: list[tuple[str, str, str, str]] = []
    for segment in _split_elective_segments(line):
        for subject, instructor, room, notes in _split_single_line_lesson(segment):
            if not subject:
                continue
            if _is_elective_header(subject) and not _strip_elective_header(subject):
                continue
            raw_items.append((subject, instructor, room, notes))
    if not raw_items:
        return []

    bare = next(
        (title for subject, *_rest in raw_items if (title := _bare_course_title(subject))),
        "",
    )
    instructors = [instructor for _s, instructor, _r, _n in raw_items if instructor]
    shared_instructor = instructors[0] if len(set(instructors)) == 1 else ""

    entries: list[tuple[str, str, str, str]] = []
    last_instructor = ""
    for subject, instructor, room, notes in raw_items:
        if not instructor:
            instructor = last_instructor or shared_instructor
        if _subject_is_metadata(subject):
            if not bare:
                continue
            subject = f"{bare} {subject}".strip()
        elif not _is_usable_subject(subject):
            continue
        subject, notes = _keep_group_markers_on_subject(subject, notes)
        if instructor:
            last_instructor = instructor
        entries.append(
            (subject, instructor, room, (notes + " (по выбору)").strip())
        )
    return entries


def _parse_cell_text(raw: str) -> list[tuple[str, str, str, str]]:
    """Extract subject, instructor, room, and extra notes from a cell string.

    The cells typically look like:
        Математический анализ,
        Зубкова С.К.
        1306

    or multi-line electives:
        Дисциплины по выбору:
        Subject 1, Instructor 1 in 1306
        Subject 2, Instructor 2 (вебинар)
    """
    if not raw or not raw.strip():
        return []

    lines = [ln.strip() for ln in raw.strip().splitlines() if ln.strip()]
    if not lines:
        return []

    if _is_elective_header(lines[0]):
        rest_first = _strip_elective_header(lines[0])
        item_lines = ([rest_first] if rest_first else []) + lines[1:]
        entries: list[tuple[str, str, str, str]] = []
        for line in item_lines:
            if _is_elective_header(line) and not _strip_elective_header(line):
                continue
            entries.extend(_parse_elective_line(line))
        return entries

    if len(lines) == 1:
        return _split_single_line_lesson(lines[0])

    return [_parse_lesson_block(lines)]


def _split_single_line_lesson(line: str) -> list[tuple[str, str, str, str]]:
    """Split a single line that might contain subject, instructor, and room."""
    # Example: "Технологии разработки ПО (Android), Зарипова Д.И. (вебинары)"
    # Example: "Проектирование веб-интерфейсов, Якушенкова А.Д. в 1305"
    
    instr_m = _INSTR_RE.search(line)

    subject = line
    instructor = ""
    room = ""
    notes = ""

    if instr_m:
        instructor = instr_m.group(1).strip()
        subject = line[:instr_m.start()].strip().rstrip(",")
        remainder = line[instr_m.end():].strip().lstrip(",").strip()

        room_m = _ROOM_RE.search(remainder)
        if room_m:
            room = room_m.group(1)
            notes = (remainder[:room_m.start()] + remainder[room_m.end():]).strip(" ,;")
        else:
            notes = remainder
    else:
        room_m = _ROOM_RE.search(line)
        if room_m:
            room = room_m.group(1)
            subject = line[:room_m.start()].strip().rstrip(",")
            notes = line[room_m.end():].strip(" ,;")

    subject = subject.strip().rstrip(",").strip()
    # A room inside an auditorium marker like "(ауд.1405)" leaves a dangling
    # "(" on the subject and ")." in the notes — drop the unbalanced tail.
    if subject.count("(") > subject.count(")"):
        subject = subject[: subject.rfind("(")].strip().rstrip(",").strip()
    if notes and not notes.strip(" ().,;"):
        notes = ""
    return [(subject, instructor.strip(), room.strip(), notes.strip())]


def _parse_lesson_block(lines: list[str]) -> tuple[str, str, str, str]:
    """Parse a multi-line lesson block into subject / instructor / room / notes."""
    subject, instructor, room, first_notes = _split_single_line_lesson(lines[0])[0]
    notes_parts: list[str] = [first_notes] if first_notes else []

    for line in lines[1:]:
        clean = line.rstrip(",").strip()
        if not clean:
            continue
        if re.match(r"^(?:ауд\.?\s*)?\d{3,4}(-\d{3,4})?$", clean, re.IGNORECASE):
            room_m = _ROOM_RE.search(clean)
            if room_m:
                room = room_m.group(1)
            continue
        if _INSTR_RE.search(clean):
            instr_match = re.match(
                r"(.*?[А-ЯЁа-яё]+\s+[А-ЯЁ]\.[А-ЯЁ]?\.?)(.*)", clean
            )
            if instr_match:
                if not instructor:
                    instructor = instr_match.group(1).strip()
                remainder = instr_match.group(2).strip()
                if remainder:
                    room_m = _ROOM_RE.search(remainder)
                    if room_m:
                        room = room_m.group(1)
                    notes_parts.append(remainder)
            elif not instructor:
                instructor = clean
            continue
        room_m = _ROOM_RE.search(clean)
        if room_m:
            room = room_m.group(1)
            leftover = (clean[:room_m.start()] + clean[room_m.end():]).strip(" ,;")
            if leftover and not re.match(r"^ауд\.?$", leftover, re.IGNORECASE):
                notes_parts.append(leftover)
        else:
            notes_parts.append(clean)

    notes = "; ".join(p for p in notes_parts if p)
    return (subject, instructor, room, notes)


def _detect_lesson_type(subject: str, instructor: str, notes: str, is_shared: bool = True) -> str:
    """Heuristic to detect lesson type from text fields."""
    text = f"{subject} {instructor} {notes}".lower()

    if any(k in text for k in ["лекци", " лек.", " лек ", "(лек."]):
        return "Лекц"
    if any(k in text for k in ["практик", "прак", " пр.", " пр ", "(пр."]):
        return "Прак"

    # Fallback based on sharing:
    # Shared sessions (cross-group) are usually lectures.
    # Non-shared sessions (per-group) are usually practices.
    if is_shared:
        return "Лекц"
    else:
        return "Прак"


def _detect_lesson_weeks(subject: str, instructor: str, notes: str) -> str:
    """Detect even/odd parity or an explicit week range like '1-9 нед'."""
    text = f"{subject} {instructor} {notes}"
    low = text.lower()

    from_to_m = re.search(r"с\s+(\d{1,2})\s+по\s+(\d{1,2})\s*нед", low)
    if from_to_m:
        start, end = int(from_to_m.group(1)), int(from_to_m.group(2))
        if 1 <= start <= end <= 22:
            return f"{start}-{end}"

    range_m = re.search(r"(\d{1,2})\s*-\s*(\d{1,2})\s*нед", low)
    if range_m:
        start, end = int(range_m.group(1)), int(range_m.group(2))
        if 1 <= start <= end <= 22:
            return f"{start}-{end}"

    from_m = re.search(r"с\s+(\d{1,2})\s*нед", low)
    if from_m:
        start = int(from_m.group(1))
        if 1 <= start <= 22:
            return f"{start}-18"

    only_m = re.search(r"(?:^|[^\d])(\d{1,2})\s*нед", low)
    if only_m and "с " not in low[max(0, only_m.start() - 3):only_m.start() + 1]:
        # "лек. 9 нед" usually means the first 9 weeks, not only week 9.
        n = int(only_m.group(1))
        if 2 <= n <= 18:
            return f"1-{n}"

    if any(k in low for k in ["нечетн", "нечет", "неч."]):
        return "odd"
    # Avoid matching "четв" (Thursday) / "четвёрг".
    if re.search(r"\bчетн|\bчет\.|\bч\.н\b|\bчн\b", low):
        return "even"

    return "all"


def _detect_start_from(subject: str, instructor: str, notes: str) -> str:
    """Parse 'с 8 сентября' into an MM-DD string, or empty."""
    text = f"{subject} {instructor} {notes}".lower()
    m = re.search(
        r"с\s+(\d{1,2})\s+(" + "|".join(_MONTH_NAMES) + r")",
        text,
    )
    if not m:
        return ""
    day = int(m.group(1))
    month = _MONTH_NAMES[m.group(2)]
    return f"{month:02d}-{day:02d}"


def parse_schedule(xlsx_bytes: bytes) -> dict[str, list[Lesson]]:
    """Parse the full schedule XLSX into a mapping of group_id → [Lesson].

    Args:
        xlsx_bytes: Raw XLSX bytes exported from Google Sheets.

    Returns:
        Dictionary mapping group identifiers (e.g. "11-501") to their
        list of Lesson objects.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # 1. Expand merged cells so every cell in a merge range has the value
    #    and hyperlink of the top-left cell.
    #    We build a dense in-memory grid: grid[row_idx][col_idx] = (str_value, link)
    #    using 0-based indexing for compatibility with previous logic.
    rows_data: list[list[tuple[str, str]]] = []
    # Determine max row/col
    max_row = ws.max_row
    max_col = ws.max_column

    # Pre-fill grid with empty strings
    # Note: openpyxl rows/cols are 1-based, list is 0-based.
    # ws.cell(r, c) -> grid[r-1][c-1]
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            val = str(cell.value) if cell.value is not None else ""
            link = ""
            if cell.hyperlink and cell.hyperlink.target:
                link = cell.hyperlink.target
            row_vals.append((val, link))
        rows_data.append(row_vals)

    # Apply merges
    for merge_range in ws.merged_cells.ranges:
        # merge_range boundaries are inclusive and 1-based
        min_col, min_row, max_col_rng, max_row_rng = (
            merge_range.min_col,
            merge_range.min_row,
            merge_range.max_col,
            merge_range.max_row,
        )
        
        # Get top-left value
        # 0-based indices
        tl_val = rows_data[min_row - 1][min_col - 1]
        
        # Fill the range
        for r in range(min_row - 1, max_row_rng):
            for c in range(min_col - 1, max_col_rng):
                rows_data[r][c] = tl_val

    # 2. Existing parsing logic using the dense grid
    if len(rows_data) < 3:
        raise ValueError("XLSX has too few rows to contain a schedule")

    # --- Extract group IDs from row index 1 (0-based) -----------------
    group_row = rows_data[1]
    groups: dict[int, str] = {}
    for col_idx in range(2, len(group_row)):
        # group_row contains (value, link) tuples now
        gid = group_row[col_idx][0].strip()
        if gid:
            groups[col_idx] = gid

    # --- Walk through data rows, tracking current day ------------------
    schedule: dict[str, list[Lesson]] = {gid: [] for gid in groups.values()}
    current_day: int | None = None

    for row in rows_data[2:]:
        if not row:
            continue

        # Check if column A contains a day-of-week marker.
        # row[0] is (val, link)
        col_a_val = row[0][0] if len(row) > 0 else ""
        day_match = _DAY_RE.search(col_a_val)
        if day_match:
            matched_text = _normalize_day(day_match.group(1))
            for day_name, day_idx in DAY_MARKERS.items():
                if matched_text == day_name:
                    current_day = day_idx
                    break

        if current_day is None:
            continue

        # Check if column B contains a time slot.
        col_b_val = row[1][0] if len(row) > 1 else ""
        time_match = _TIME_RE.search(col_b_val)
        if not time_match:
            continue

        time_start = f"{time_match.group(1)}:{time_match.group(2)}"
        time_end = f"{time_match.group(3)}:{time_match.group(4)}"

        # Parse each group column.
        for col_idx, gid in groups.items():
            if col_idx >= len(row):
                continue
            
            cell_val, cell_link = row[col_idx]
            cell_val = cell_val.strip()
            
            if not cell_val:
                continue

            # Check if this cell's content is shared with ANY other group in this row
            is_shared = False
            for other_col_idx, other_gid in groups.items():
                if other_col_idx == col_idx:
                    continue
                if other_col_idx < len(row):
                    other_val = row[other_col_idx][0].strip()
                    if other_val == cell_val:
                        is_shared = True
                        break

            results = [
                (subject.strip().rstrip(":").strip(), instructor, room, notes)
                for subject, instructor, room, notes in _parse_cell_text(cell_val)
            ]
            usable = [item for item in results if _is_usable_subject(item[0])]
            n_entries = len(usable)
            for subject, instructor, room, notes in usable:

                # Keep notes exactly as in the source calendar cell.
                raw_notes = cell_val
                # Week/start markers are per-lesson (the parsed line), not the
                # whole cell — elective cells mix several options. Same for
                # the type: the whole cell leaks e.g. "прак." from a sibling
                # practice into a shared lecture (Психология управления 1-9
                # нед. Пучкова vs с 10 нед. прак. гр.№1 Зайнуллин).
                lesson_type = _detect_lesson_type(subject, instructor, notes, is_shared)
                weeks = _detect_lesson_weeks(subject, instructor, notes)
                start_from = _detect_start_from(subject, instructor, notes)

                # A merged cell has at most one hyperlink. Copying it onto
                # every elective in the cell attaches e.g. a robotics webinar
                # to an in-person blockchain class in 1304. Single-entry cells
                # can keep the hyperlink; the links tab fills the rest.
                lesson_link = cell_link if n_entries == 1 else ""

                schedule[gid].append(
                    Lesson(
                        day=current_day,
                        time_start=time_start,
                        time_end=time_end,
                        subject=subject,
                        instructor=instructor,
                        room=room,
                        notes=raw_notes,
                        link=lesson_link,
                        type=lesson_type,
                        weeks=weeks,
                        start_from=start_from,
                    )
                )

    for gid, lessons in schedule.items():
        seen: set[tuple] = set()
        unique: list[Lesson] = []
        for lesson in lessons:
            key = (
                lesson.day,
                lesson.time_start,
                lesson.time_end,
                lesson.subject.casefold(),
                lesson.instructor.casefold(),
                lesson.room,
                lesson.weeks,
            )
            if key in seen:
                continue
            seen.add(key)
            unique.append(lesson)
        schedule[gid] = unique

    return schedule
